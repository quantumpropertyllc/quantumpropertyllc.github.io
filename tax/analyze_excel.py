
import openpyxl
import sys

file_path = '/Volumes/HDD/iOS/2025TaxHtml/25_1040.xlsx'

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print("Sheet names:", wb.sheetnames)
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n--- Sheet: {sheet} ---")
        # Print first 10 rows to get a feel
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 10: break
            print(row)
            
except Exception as e:
    print(f"Error: {e}")
